from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

logger = logging.getLogger(__name__)


def _find_column_index(ws, column_name: str) -> int | None:
    """헤더 행에서 컬럼 인덱스(1-based) 반환."""
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and column_name in str(cell.value).strip():
            return col_idx
    return None


def mark_sent_items(
    excel_path: str | Path,
    sheet_name: str,
    row_indices: list[int],
    column_name: str = "갱신신청기한",
    border_color: str = "FF0000",
    border_style: str = "thin",
) -> None:
    """발송 완료 품목의 갱신신청기한 셀에 붉은 테두리 적용."""
    path = Path(excel_path)
    if not path.exists():
        logger.warning("엑셀 파일 없음: %s", path)
        return

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        logger.warning("시트 '%s' 없음 (파일: %s)", sheet_name, path)
        wb.close()
        return

    ws = wb[sheet_name]
    col_idx = _find_column_index(ws, column_name)
    if col_idx is None:
        logger.warning("컬럼 '%s' 없음 (시트: %s)", column_name, sheet_name)
        wb.close()
        return

    red_border = Border(
        left=Side(style=border_style, color=border_color),
        right=Side(style=border_style, color=border_color),
        top=Side(style=border_style, color=border_color),
        bottom=Side(style=border_style, color=border_color),
    )

    # row_indices는 데이터 행 번호 (1-based, 헤더 제외) → 엑셀 행 = row_index + 1
    for row_idx in row_indices:
        excel_row = row_idx + 1
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.border = red_border

    wb.save(path)
    wb.close()
    logger.info(
        "%s 시트에서 %d건 붉은 테두리 적용 완료", sheet_name, len(row_indices)
    )
